from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
import json

def index(request):
    return render(request, 'index.html')

def export_excel(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.POST.get('exportData'))

            workbook = Workbook()
            sheet = workbook.active

            for table_data in data['tables']:
                print_value = table_data['PRINT']
                new_task = table_data['NEW_TASK']

                for row_data in table_data['tableData']:
                    sheet.append(row_data)

                # Add a 2-row blank gap without borders
                for _ in range(0):
                    sheet.append([])

                # Add borders and center text in cells
                row_i = 1
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell.border = Border(left=Side(border_style='thin'),
                                                 right=Side(border_style='thin'),
                                                 top=Side(border_style='thin'),
                                                 bottom=Side(border_style='thin'))
                            if cell.value in ['CWI', 'SPLICE DETAILS']:
                                sheet.merge_cells(start_row=row_i, end_row=row_i + 1, start_column=6, end_column=9)
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                    row_i += 1

                # Set column width based on content or default to 20px
                for column in sheet.columns:
                    max_length = 0
                    has_value = any(cell.value for cell in column)

                    if has_value:
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass

                        # Set the width based on the content
                        adjusted_width = max_length + 2
                    else:
                        # Default width of 20 pixels for columns with no content
                        adjusted_width = 10

                    try:
                        sheet.column_dimensions[column[0].column_letter].width = adjusted_width
                    except:
                        print("Error setting column width")

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Schema_tables.xlsx'
            workbook.save(response)

            return response
        except Exception as e:
            # Handle exceptions, log them, or return an appropriate error response
            return JsonResponse({'error': str(e)}, status=500)
    else:
        # Return an appropriate error response for non-POST requests
        return JsonResponse({'error': 'Invalid request method'}, status=400)
